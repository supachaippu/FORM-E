import os
import re
import csv
import pandas as pd
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_part_no(part):
    if not part or pd.isna(part):
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(part)).upper()

def clean_eng_name(name):
    if not name or pd.isna(name):
        return ""
    s = str(name).upper()
    # Remove hyphens followed by space/newline, then remove all non-alphanumeric
    s = s.replace('- ', '').replace('-\n', '').replace('-', '')
    return re.sub(r'[^A-Z0-9]', '', s)

def extract_fields_from_pdf(pdf_desc, excel_part_no):
    # Strip package prefix (e.g. "ONE (1) PACKAGE OF ")
    prefix_match = re.match(r'^(?:[A-Z\s\-]+ \(\d+\) PACKAGES? OF\s+)', pdf_desc, re.IGNORECASE)
    if prefix_match:
        prefix = prefix_match.group(0)
        remaining = pdf_desc[len(prefix):].strip()
    else:
        remaining = pdf_desc.strip()
        
    # Extract HS Code
    hs_code = ""
    hs_match = re.search(r'HS\s*CODE:\s*([0-9\.]+)', remaining, re.IGNORECASE)
    if hs_match:
        hs_code = hs_match.group(1).strip()
        remaining = remaining[:hs_match.start()].strip()
        
    # Extract Part Number and English Name
    clean_ex = clean_part_no(excel_part_no)
    
    alpha_indices = [i for i, c in enumerate(remaining) if c.isalnum()]
    alpha_chars = ''.join([remaining[i].upper() for i in alpha_indices])
    
    match_idx = alpha_chars.find(clean_ex)
    if match_idx != -1:
        start_in_remaining = alpha_indices[match_idx]
        end_in_remaining = alpha_indices[match_idx + len(clean_ex) - 1] + 1
        
        eng_name = remaining[:start_in_remaining].strip()
        pdf_part = remaining[start_in_remaining:end_in_remaining].strip()
        
        # Clean trailing spacing/punctuation from English name
        eng_name = re.sub(r'[\s\-\,]+$', '', eng_name)
        return eng_name, pdf_part, hs_code
    else:
        # Fallback: split by multiple spaces
        parts = re.split(r'\s{2,}', remaining)
        if len(parts) >= 2:
            return parts[0], parts[1], hs_code
        else:
            return remaining, "", hs_code

def extract_pdf_items(pdf_path):
    doc = fitz.open(pdf_path)
    pdf_items = []
    current_item = None
    
    for page_idx in range(len(doc)):
        page_num = page_idx + 1
        page = doc[page_idx]
        blocks = page.get_text("blocks")
        
        page_items = []
        page_qtys = []
        page_descs = []
        
        for b in blocks:
            x0, y0, x1, y1, text, block_no, block_type = b
            text = text.strip()
            if not text:
                continue
                
            x0 = round(x0, 1)
            y0 = round(y0, 1)
            x1 = round(x1, 1)
            y1 = round(y1, 1)
            
            # Item number (typically x0 around 65-85, y0 in the table area)
            if 65 <= x0 <= 85 and 350 <= y0 <= 600:
                lines = [l.strip() for l in text.split('\n') if l.strip()]
                for line in lines:
                    try:
                        val = int(line)
                        page_items.append((y0, val))
                    except ValueError:
                        pass
                        
            # Description column (x0 around 160-180)
            elif 160 <= x0 <= 180 and 350 <= y0 <= 600:
                page_descs.append((y0, text))
                
            # Quantity column (x0 around 420-470)
            elif 420 <= x0 <= 470 and 350 <= y0 <= 600:
                page_qtys.append((y0, text))
                
        # Sort by y coordinate
        page_items.sort(key=lambda x: x[0])
        page_qtys.sort(key=lambda x: x[0])
        page_descs.sort(key=lambda x: x[0])
        
        # Process continuation from previous page
        if page_descs:
            first_item_y = page_items[0][0] if page_items else 999.0
            first_desc_y, first_desc_text = page_descs[0]
            
            if first_desc_y < first_item_y and first_desc_y < 390:
                # Skip TRUCK SPARE PARTS header on page 1
                if page_num == 1 and first_desc_text == 'TRUCK SPARE PARTS':
                    page_descs.pop(0)
                else:
                    if current_item is not None:
                        current_item['descs'].append(first_desc_text)
                        page_descs.pop(0)
                    else:
                        page_descs.pop(0)
                        
        # Match items, descs, and qtys on this page
        for y_item, item_no in page_items:
            # Find closest qty
            matched_qty = ""
            best_qty_diff = 999.0
            best_qty_idx = -1
            for idx, (y_qty, qty_text) in enumerate(page_qtys):
                diff = abs(y_qty - y_item)
                if diff < best_qty_diff:
                    best_qty_diff = diff
                    best_qty_idx = idx
            if best_qty_idx != -1 and best_qty_diff < 10.0:
                matched_qty = page_qtys[best_qty_idx][1]
                page_qtys.pop(best_qty_idx)
                
            # Find closest desc
            matched_desc = ""
            best_desc_diff = 999.0
            best_desc_idx = -1
            for idx, (y_desc, desc_text) in enumerate(page_descs):
                diff = abs(y_desc - y_item)
                if diff < best_desc_diff:
                    best_desc_diff = diff
                    best_desc_idx = idx
            if best_desc_idx != -1 and best_desc_diff < 10.0:
                matched_desc = page_descs[best_desc_idx][1]
                page_descs.pop(best_desc_idx)
                
            current_item = {
                'item_no': item_no,
                'qty_raw': matched_qty,
                'descs': [matched_desc] if matched_desc else [],
                'page': page_num
            }
            pdf_items.append(current_item)
            
    # Post-process extracted PDF items
    processed_items = []
    for item in pdf_items:
        full_desc = " ".join(item['descs']).replace('\n', ' ')
        
        # Parse Qty
        qty_val = 0
        qty_num_match = re.search(r'(\d+)', item['qty_raw'])
        if qty_num_match:
            qty_val = int(qty_num_match.group(1))
            
        processed_items.append({
            'item_no': item['item_no'],
            'qty_raw': item['qty_raw'],
            'qty_val': qty_val,
            'full_desc': full_desc,
            'page': item['page']
        })
        
    return processed_items

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, "Invoice and Packing list 326T1.xlsx")
    pdf_path = os.path.join(base_dir, "FE-T1.pdf")
    output_path = os.path.join(base_dir, "form_e_comparison_report.csv")
    
    print(f"Reading Excel: {excel_path}")
    df_excel = pd.read_excel(excel_path, sheet_name="INVOICE FOR CNY")
    
    excel_items = []
    for idx, row in df_excel.iterrows():
        try:
            sl = int(row.iloc[0])
            part_no = str(row.iloc[5]).strip()
            ocr_no = str(row.iloc[6]).strip()
            qty = int(row.iloc[7])
            hs_code = str(row.iloc[11]).strip()
            thai_name = str(row.iloc[2]).strip()
            eng_name = str(row.iloc[4]).strip()
            
            excel_items.append({
                'sl': sl,
                'part_no': part_no,
                'ocr_no': ocr_no,
                'qty': qty,
                'hs_code': hs_code,
                'thai_name': thai_name,
                'eng_name': eng_name
            })
        except:
            continue
            
    print(f"Loaded {len(excel_items)} items from Excel sheet.")
    
    print(f"Parsing PDF: {pdf_path}")
    pdf_items = extract_pdf_items(pdf_path)
    print(f"Extracted {len(pdf_items)} items from PDF.")
    
    report_rows = []
    mismatches_count = 0
    
    max_len = max(len(excel_items), len(pdf_items))
    for i in range(max_len):
        ex = excel_items[i] if i < len(excel_items) else None
        pdf = pdf_items[i] if i < len(pdf_items) else None
        
        status = "MATCH"
        mismatch_reasons = []
        
        if ex and pdf:
            # Parse fields from PDF using Excel Part No as key
            pdf_eng, pdf_part, pdf_hs = extract_fields_from_pdf(pdf['full_desc'], ex['part_no'])
            
            # 1. Check English name
            clean_ex_eng = clean_eng_name(ex['eng_name'])
            clean_pdf_eng = clean_eng_name(pdf_eng)
            name_match = (clean_ex_eng == clean_pdf_eng)
            
            # 2. Check Part number
            clean_ex_part = clean_part_no(ex['part_no'])
            clean_pdf_part = clean_part_no(pdf_part)
            part_match = (clean_ex_part == clean_pdf_part)
            
            # 3. Check Qty
            qty_match = (ex['qty'] == pdf['qty_val'])
            
            # 4. Check HS code (first 6 digits)
            clean_ex_hs = re.sub(r'[^0-9]', '', ex['hs_code'])[:6]
            clean_pdf_hs = re.sub(r'[^0-9]', '', pdf_hs)[:6]
            hs_match = (clean_ex_hs == clean_pdf_hs)
            
            if not name_match:
                status = "MISMATCH"
                mismatch_reasons.append("English Name mismatch")
            if not part_match:
                status = "MISMATCH"
                mismatch_reasons.append("Part No mismatch")
            if not qty_match:
                status = "MISMATCH"
                mismatch_reasons.append("Qty mismatch")
            if not hs_match:
                status = "MISMATCH"
                mismatch_reasons.append("HS Code mismatch")
                
            report_rows.append({
                'Item No (Excel)': ex['sl'],
                'Item No (PDF)': pdf['item_no'],
                'PDF Page': pdf['page'],
                'English Name (Excel)': ex['eng_name'],
                'English Name (PDF)': pdf_eng,
                'English Name Match': "MATCH" if name_match else "MISMATCH",
                'Part No (Excel)': ex['part_no'],
                'Part No (PDF)': pdf_part,
                'Part No Match': "MATCH" if part_match else "MISMATCH",
                'Qty (Excel)': ex['qty'],
                'Qty (PDF)': pdf['qty_raw'],
                'Qty Match': "MATCH" if qty_match else "MISMATCH",
                'HS Code (Excel)': ex['hs_code'],
                'HS Code (PDF)': pdf_hs,
                'HS Code Match': "MATCH" if hs_match else "MISMATCH",
                'Overall Status': status,
                'Details': "; ".join(mismatch_reasons) if mismatch_reasons else "All matched"
            })
            
            if status == "MISMATCH":
                mismatches_count += 1
                
        elif ex:
            report_rows.append({
                'Item No (Excel)': ex['sl'], 'Item No (PDF)': "", 'PDF Page': "",
                'English Name (Excel)': ex['eng_name'], 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                'Part No (Excel)': ex['part_no'], 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                'Qty (Excel)': ex['qty'], 'Qty (PDF)': "", 'Qty Match': "MISMATCH",
                'HS Code (Excel)': ex['hs_code'], 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                'Overall Status': "MISMATCH", 'Details': "Item missing in PDF"
            })
            mismatches_count += 1
        elif pdf:
            report_rows.append({
                'Item No (Excel)': "", 'Item No (PDF)': pdf['item_no'], 'PDF Page': pdf['page'],
                'English Name (Excel)': "", 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                'Part No (Excel)': "", 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                'Qty (Excel)': "", 'Qty (PDF)': pdf['qty_raw'], 'Qty Match': "MISMATCH",
                'HS Code (Excel)': "", 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                'Overall Status': "MISMATCH", 'Details': "Item missing in Excel"
            })
            mismatches_count += 1

    # Write CSV report
    headers = [
        'Item No (Excel)', 'Item No (PDF)', 'PDF Page', 
        'English Name (Excel)', 'English Name (PDF)', 'English Name Match',
        'Part No (Excel)', 'Part No (PDF)', 'Part No Match',
        'Qty (Excel)', 'Qty (PDF)', 'Qty Match',
        'HS Code (Excel)', 'HS Code (PDF)', 'HS Code Match',
        'Overall Status', 'Details'
    ]
    
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(report_rows)
    print(f"Report written to: {output_path}")

    # Write Styled Excel report
    xlsx_path = output_path.replace(".csv", ".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Report"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    
    regular_font = Font(name=font_family, size=11)
    
    # Matching coloring (green)
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    match_font = Font(name=font_family, size=11, color="375623")
    
    # Mismatch coloring (red)
    mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    mismatch_font = Font(name=font_family, size=11, color="C00000", bold=True)
    
    # Overall Status formatting
    status_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    status_match_font = Font(name=font_family, size=11, color="006100", bold=True)
    
    status_mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    status_mismatch_font = Font(name=font_family, size=11, color="9C0006", bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    
    for r_idx, row_dict in enumerate(report_rows, start=2):
        row_values = [row_dict[h] for h in headers]
        ws.append(row_values)
        ws.row_dimensions[r_idx].height = 20
        
        # We want to format cells side by side.
        # Header indexes:
        # Col 1: Item No (Excel)
        # Col 2: Item No (PDF)
        # Col 3: PDF Page
        # Col 4: English Name (Excel)
        # Col 5: English Name (PDF)
        # Col 6: English Name Match
        # Col 7: Part No (Excel)
        # Col 8: Part No (PDF)
        # Col 9: Part No Match
        # Col 10: Qty (Excel)
        # Col 11: Qty (PDF)
        # Col 12: Qty Match
        # Col 13: HS Code (Excel)
        # Col 14: HS Code (PDF)
        # Col 15: HS Code Match
        # Col 16: Overall Status
        # Col 17: Details
        
        # Clean Excel row items
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border
            
            # Basic alignments
            if h in ['Item No (Excel)', 'Item No (PDF)', 'PDF Page', 'Qty (Excel)', 'Qty (PDF)', 'HS Code (Excel)', 'HS Code (PDF)', 'English Name Match', 'Part No Match', 'Qty Match', 'HS Code Match', 'Overall Status']:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        
        # Apply conditional formats for English Name Match
        name_match_status = row_dict['English Name Match']
        fill_to_use = match_fill if name_match_status == "MATCH" else mismatch_fill
        font_to_use = match_font if name_match_status == "MATCH" else mismatch_font
        ws.cell(row=r_idx, column=4).fill = fill_to_use    # English Name (Excel)
        ws.cell(row=r_idx, column=4).font = font_to_use
        ws.cell(row=r_idx, column=5).fill = fill_to_use    # English Name (PDF)
        ws.cell(row=r_idx, column=5).font = font_to_use
        ws.cell(row=r_idx, column=6).fill = fill_to_use    # Match column
        ws.cell(row=r_idx, column=6).font = font_to_use

        # Apply conditional formats for Part No Match
        part_match_status = row_dict['Part No Match']
        fill_to_use = match_fill if part_match_status == "MATCH" else mismatch_fill
        font_to_use = match_font if part_match_status == "MATCH" else mismatch_font
        ws.cell(row=r_idx, column=7).fill = fill_to_use    # Part No (Excel)
        ws.cell(row=r_idx, column=7).font = font_to_use
        ws.cell(row=r_idx, column=8).fill = fill_to_use    # Part No (PDF)
        ws.cell(row=r_idx, column=8).font = font_to_use
        ws.cell(row=r_idx, column=9).fill = fill_to_use    # Match column
        ws.cell(row=r_idx, column=9).font = font_to_use

        # Apply conditional formats for Qty Match
        qty_match_status = row_dict['Qty Match']
        fill_to_use = match_fill if qty_match_status == "MATCH" else mismatch_fill
        font_to_use = match_font if qty_match_status == "MATCH" else mismatch_font
        ws.cell(row=r_idx, column=10).fill = fill_to_use   # Qty (Excel)
        ws.cell(row=r_idx, column=10).font = font_to_use
        ws.cell(row=r_idx, column=11).fill = fill_to_use   # Qty (PDF)
        ws.cell(row=r_idx, column=11).font = font_to_use
        ws.cell(row=r_idx, column=12).fill = fill_to_use   # Match column
        ws.cell(row=r_idx, column=12).font = font_to_use

        # Apply conditional formats for HS Code Match
        hs_match_status = row_dict['HS Code Match']
        fill_to_use = match_fill if hs_match_status == "MATCH" else mismatch_fill
        font_to_use = match_font if hs_match_status == "MATCH" else mismatch_font
        ws.cell(row=r_idx, column=13).fill = fill_to_use   # HS Code (Excel)
        ws.cell(row=r_idx, column=13).font = font_to_use
        ws.cell(row=r_idx, column=14).fill = fill_to_use   # HS Code (PDF)
        ws.cell(row=r_idx, column=14).font = font_to_use
        ws.cell(row=r_idx, column=15).fill = fill_to_use   # Match column
        ws.cell(row=r_idx, column=15).font = font_to_use

        # Overall status formatting
        overall_status = row_dict['Overall Status']
        status_cell = ws.cell(row=r_idx, column=16)
        if overall_status == "MATCH":
            status_cell.fill = status_match_fill
            status_cell.font = status_match_font
        else:
            status_cell.fill = status_mismatch_fill
            status_cell.font = status_mismatch_font
            
    # Auto-adjust column widths
    for col in ws.columns:
        col_max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > col_max_len:
                    col_max_len = len(line)
        ws.column_dimensions[col_letter].width = max(min(col_max_len + 3, 50), 10)
        
    wb.save(xlsx_path)
    print(f"Styled Excel written to: {xlsx_path}")
    print(f"Total Items: {max_len}")
    print(f"Mismatches: {mismatches_count}")
    print("Execution complete.")

if __name__ == "__main__":
    main()
