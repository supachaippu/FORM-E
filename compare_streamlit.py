import io
import re
import csv
import pandas as pd
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# Configure Streamlit Page
st.set_page_config(
    page_title="FORM E - Comparative Verification Tool",
    page_icon="📋",
    layout="wide"
)

# Custom Premium Styling
st.markdown("""
<style>
    /* Main container styling */
    .reportview-container {
        background-color: #F8F9FA;
    }
    
    /* Header section styling */
    .header-container {
        background: linear-gradient(135deg, #1f4068, #162447);
        padding: 30px;
        border-radius: 12px;
        color: white;
        margin-bottom: 25px;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
    }
    
    .header-title {
        font-size: 32px;
        font-weight: 700;
        margin-bottom: 10px;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    .header-subtitle {
        font-size: 16px;
        font-style: italic;
        opacity: 0.9;
    }
    
    /* Metric Card Styling */
    .metric-card {
        background: white;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border-top: 5px solid #365F91;
        text-align: center;
    }
    
    .metric-val {
        font-size: 36px;
        font-weight: 800;
        color: #2c3e50;
    }
</style>
""", unsafe_allow_html=True)

# Helper functions for comparison
def clean_part_no(part):
    if not part or pd.isna(part):
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(part)).upper()

def clean_eng_name(name):
    if not name or pd.isna(name):
        return ""
    s = str(name).upper()
    s = s.replace('- ', '').replace('-\n', '').replace('-', '')
    return re.sub(r'[^A-Z0-9]', '', s)

def extract_fields_from_pdf(pdf_desc, excel_part_no):
    # Strip package prefix
    prefix_match = re.match(r'^(?:[A-Z\s\-]+ \(\d+\) PACKAGES? OF\s+)', pdf_desc, re.IGNORECASE)
    if prefix_match:
        prefix = prefix_match.group(0)
        remaining = pdf_desc[len(prefix):].strip()
    else:
        remaining = pdf_desc.strip()
        
    # Extract HS Code
    hs_code = ""
    hs_match = re.search(r'HS\s*CODE:\s*([0-9\.]+)', remaining, re.IGNORECASE)
    if hs_match:
        hs_code = hs_match.group(1).strip()
        remaining = remaining[:hs_match.start()].strip()
        
    # Extract Part Number and English Name
    clean_ex = clean_part_no(excel_part_no)
    
    alpha_indices = [i for i, c in enumerate(remaining) if c.isalnum()]
    alpha_chars = ''.join([remaining[i].upper() for i in alpha_indices])
    
    match_idx = alpha_chars.find(clean_ex)
    if match_idx != -1:
        start_in_remaining = alpha_indices[match_idx]
        end_in_remaining = alpha_indices[match_idx + len(clean_ex) - 1] + 1
        
        eng_name = remaining[:start_in_remaining].strip()
        pdf_part = remaining[start_in_remaining:end_in_remaining].strip()
        
        eng_name = re.sub(r'[\s\-\,]+$', '', eng_name)
        return eng_name, pdf_part, hs_code
    else:
        parts = re.split(r'\s{2,}', remaining)
        if len(parts) >= 2:
            return parts[0], parts[1], hs_code
        else:
            return remaining, "", hs_code

def extract_pdf_items_from_stream(pdf_file):
    # Open PyMuPDF from bytes stream
    pdf_bytes = pdf_file.read()
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pdf_items = []
    current_item = None
    
    for page_idx in range(len(doc)):
        page_num = page_idx + 1
        page = doc[page_idx]
        blocks = page.get_text("blocks")
        
        page_items = []
        page_qtys = []
        page_descs = []
        
        for b in blocks:
            x0, y0, x1, y1, text, block_no, block_type = b
            text = text.strip()
            if not text:
                continue
                
            x0 = round(x0, 1)
            y0 = round(y0, 1)
            x1 = round(x1, 1)
            y1 = round(y1, 1)
            
            # Item number (typically x0 around 65-85, y0 in the table area)
            if 65 <= x0 <= 85 and 350 <= y0 <= 600:
                lines = [l.strip() for l in text.split('\n') if l.strip()]
                for line in lines:
                    try:
                        val = int(line)
                        page_items.append((y0, val))
                    except ValueError:
                        pass
                        
            # Description column (x0 around 160-180)
            elif 160 <= x0 <= 180 and 350 <= y0 <= 600:
                page_descs.append((y0, text))
                
            # Quantity column (x0 around 420-470)
            elif 420 <= x0 <= 470 and 350 <= y0 <= 600:
                page_qtys.append((y0, text))
                
        # Sort by y coordinate
        page_items.sort(key=lambda x: x[0])
        page_qtys.sort(key=lambda x: x[0])
        page_descs.sort(key=lambda x: x[0])
        
        # Process continuation from previous page
        if page_descs:
            first_item_y = page_items[0][0] if page_items else 999.0
            first_desc_y, first_desc_text = page_descs[0]
            
            if first_desc_y < first_item_y and first_desc_y < 390:
                if page_num == 1 and first_desc_text == 'TRUCK SPARE PARTS':
                    page_descs.pop(0)
                else:
                    if current_item is not None:
                        current_item['descs'].append(first_desc_text)
                        page_descs.pop(0)
                    else:
                        page_descs.pop(0)
                        
        # Match items, descs, and qtys on this page
        for y_item, item_no in page_items:
            # Find closest qty
            matched_qty = ""
            best_qty_diff = 999.0
            best_qty_idx = -1
            for idx, (y_qty, qty_text) in enumerate(page_qtys):
                diff = abs(y_qty - y_item)
                if diff < best_qty_diff:
                    best_qty_diff = diff
                    best_qty_idx = idx
            if best_qty_idx != -1 and best_qty_diff < 10.0:
                matched_qty = page_qtys[best_qty_idx][1]
                page_qtys.pop(best_qty_idx)
                
            # Find closest desc
            matched_desc = ""
            best_desc_diff = 999.0
            best_desc_idx = -1
            for idx, (y_desc, desc_text) in enumerate(page_descs):
                diff = abs(y_desc - y_item)
                if diff < best_desc_diff:
                    best_desc_diff = diff
                    best_desc_idx = idx
            if best_desc_idx != -1 and best_desc_diff < 10.0:
                matched_desc = page_descs[best_desc_idx][1]
                page_descs.pop(best_desc_idx)
                
            current_item = {
                'item_no': item_no,
                'qty_raw': matched_qty,
                'descs': [matched_desc] if matched_desc else [],
                'page': page_num
            }
            pdf_items.append(current_item)
            
    # Post-process extracted PDF items
    processed_items = []
    for item in pdf_items:
        full_desc = " ".join(item['descs']).replace('\n', ' ')
        
        # Parse Qty
        qty_val = 0
        qty_num_match = re.search(r'(\d+)', item['qty_raw'])
        if qty_num_match:
            qty_val = int(qty_num_match.group(1))
            
        processed_items.append({
            'item_no': item['item_no'],
            'qty_raw': item['qty_raw'],
            'qty_val': qty_val,
            'full_desc': full_desc,
            'page': item['page']
        })
        
    return processed_items

# Streamlit UI Header
st.markdown("""
<div class="header-container">
    <div class="header-title">📋 ระบบตรวจสอบและเปรียบเทียบเอกสาร FORM E</div>
    <div class="header-subtitle">อัปโหลดไฟล์ Excel และ PDF เพื่อเปรียบเทียบหาความแตกต่างระหว่าง English Name, Part No, Qty และ HS Code</div>
</div>
""", unsafe_allow_html=True)

# Layout: Two columns for uploaders
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. อัปโหลดไฟล์ Excel (Invoice)")
    excel_file = st.file_uploader("เลือกไฟล์ Excel (.xlsx, .xls)", type=["xlsx", "xls"])
    
with col2:
    st.subheader("2. อัปโหลดไฟล์ PDF (FORM E)")
    pdf_file = st.file_uploader("เลือกไฟล์ PDF (.pdf)", type=["pdf"])

# Options panel if excel file is uploaded
selected_sheet = "INVOICE FOR CNY"
if excel_file:
    try:
        # Load sheets dynamically
        xl = pd.ExcelFile(excel_file)
        sheet_names = xl.sheet_names
        
        st.subheader("3. การตั้งค่าแผ่นงาน (Sheet configuration)")
        default_index = 0
        if "INVOICE FOR CNY" in sheet_names:
            default_index = sheet_names.index("INVOICE FOR CNY")
        elif "INVOICE FOR USD" in sheet_names:
            default_index = sheet_names.index("INVOICE FOR USD")
            
        selected_sheet = st.selectbox(
            "เลือกแผ่นงานใน Excel ที่ต้องการเปรียบเทียบ:", 
            sheet_names, 
            index=default_index
        )
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการอ่านรายชื่อ Sheet จาก Excel: {e}")

# Run button
if excel_file and pdf_file:
    st.markdown("---")
    if st.button("🚀 เริ่มเปรียบเทียบเอกสาร (Run Comparison)", type="primary", use_container_width=True):
        with st.spinner("กำลังวิเคราะห์สกัดข้อมูลและเปรียบเทียบรายการสินค้า... กรุณารอซักครู่"):
            try:
                # Reset file streams to beginning
                excel_file.seek(0)
                pdf_file.seek(0)
                
                # Load Excel items
                df_excel = pd.read_excel(excel_file, sheet_name=selected_sheet)
                
                excel_items = []
                for idx, row in df_excel.iterrows():
                    try:
                        sl = int(row.iloc[0])
                        part_no = str(row.iloc[5]).strip()
                        ocr_no = str(row.iloc[6]).strip()
                        qty = int(row.iloc[7])
                        hs_code = str(row.iloc[11]).strip()
                        thai_name = str(row.iloc[2]).strip()
                        eng_name = str(row.iloc[4]).strip()
                        
                        excel_items.append({
                            'sl': sl, 'part_no': part_no, 'ocr_no': ocr_no, 'qty': qty,
                            'hs_code': hs_code, 'thai_name': thai_name, 'eng_name': eng_name
                        })
                    except:
                        continue
                
                # Load PDF items
                pdf_items = extract_pdf_items_from_stream(pdf_file)
                
                # Comparison logic
                report_rows = []
                mismatches_count = 0
                
                max_len = max(len(excel_items), len(pdf_items))
                for i in range(max_len):
                    ex = excel_items[i] if i < len(excel_items) else None
                    pdf = pdf_items[i] if i < len(pdf_items) else None
                    
                    status = "MATCH"
                    mismatch_reasons = []
                    
                    if ex and pdf:
                        pdf_eng, pdf_part, pdf_hs = extract_fields_from_pdf(pdf['full_desc'], ex['part_no'])
                        
                        # Check English name
                        clean_ex_eng = clean_eng_name(ex['eng_name'])
                        clean_pdf_eng = clean_eng_name(pdf_eng)
                        name_match = (clean_ex_eng == clean_pdf_eng)
                        
                        # Check Part number
                        clean_ex_part = clean_part_no(ex['part_no'])
                        clean_pdf_part = clean_part_no(pdf_part)
                        part_match = (clean_ex_part == clean_pdf_part)
                        
                        # Check Qty
                        qty_match = (ex['qty'] == pdf['qty_val'])
                        
                        # Check HS code (first 6 digits)
                        clean_ex_hs = re.sub(r'[^0-9]', '', ex['hs_code'])[:6]
                        clean_pdf_hs = re.sub(r'[^0-9]', '', pdf_hs)[:6]
                        hs_match = (clean_ex_hs == clean_pdf_hs)
                        
                        if not name_match:
                            status = "MISMATCH"
                            mismatch_reasons.append("English Name mismatch")
                        if not part_match:
                            status = "MISMATCH"
                            mismatch_reasons.append("Part No mismatch")
                        if not qty_match:
                            status = "MISMATCH"
                            mismatch_reasons.append("Qty mismatch")
                        if not hs_match:
                            status = "MISMATCH"
                            mismatch_reasons.append("HS Code mismatch")
                            
                        report_rows.append({
                            'Item No (Excel)': ex['sl'], 'Item No (PDF)': pdf['item_no'], 'PDF Page': pdf['page'],
                            'English Name (Excel)': ex['eng_name'], 'English Name (PDF)': pdf_eng,
                            'English Name Match': "MATCH" if name_match else "MISMATCH",
                            'Part No (Excel)': ex['part_no'], 'Part No (PDF)': pdf_part,
                            'Part No Match': "MATCH" if part_match else "MISMATCH",
                            'Qty (Excel)': ex['qty'], 'Qty (PDF)': pdf['qty_raw'],
                            'Qty Match': "MATCH" if qty_match else "MISMATCH",
                            'HS Code (Excel)': ex['hs_code'], 'HS Code (PDF)': pdf_hs,
                            'HS Code Match': "MATCH" if hs_match else "MISMATCH",
                            'Overall Status': status,
                            'Details': "; ".join(mismatch_reasons) if mismatch_reasons else "All matched"
                        })
                        
                        if status == "MISMATCH":
                            mismatches_count += 1
                            
                    elif ex:
                        report_rows.append({
                            'Item No (Excel)': ex['sl'], 'Item No (PDF)': "", 'PDF Page': "",
                            'English Name (Excel)': ex['eng_name'], 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                            'Part No (Excel)': ex['part_no'], 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                            'Qty (Excel)': ex['qty'], 'Qty (PDF)': "", 'Qty Match': "MISMATCH",
                            'HS Code (Excel)': ex['hs_code'], 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                            'Overall Status': "MISMATCH", 'Details': "Item missing in PDF"
                        })
                        mismatches_count += 1
                    elif pdf:
                        report_rows.append({
                            'Item No (Excel)': "", 'Item No (PDF)': pdf['item_no'], 'PDF Page': pdf['page'],
                            'English Name (Excel)': "", 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                            'Part No (Excel)': "", 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                            'Qty (Excel)': "", 'Qty (PDF)': pdf['qty_raw'], 'Qty Match': "MISMATCH",
                            'HS Code (Excel)': "", 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                            'Overall Status': "MISMATCH", 'Details': "Item missing in Excel"
                        })
                        mismatches_count += 1
                
                # Make Dataframe for UI
                df_report = pd.DataFrame(report_rows)
                
                # ------------------- Write styled Excel in memory -------------------
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Comparison Report"
                ws.views.sheetView[0].showGridLines = True
                
                font_family = "Segoe UI"
                header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
                regular_font = Font(name=font_family, size=11)
                
                match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                match_font = Font(name=font_family, size=11, color="375623")
                mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                mismatch_font = Font(name=font_family, size=11, color="C00000", bold=True)
                
                status_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_match_font = Font(name=font_family, size=11, color="006100", bold=True)
                status_mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_mismatch_font = Font(name=font_family, size=11, color="9C0006", bold=True)
                
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")
                thin_side = Side(border_style="thin", color="D9D9D9")
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                headers = list(df_report.columns)
                ws.append(headers)
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                ws.row_dimensions[1].height = 28
                
                for r_idx, row_dict in enumerate(report_rows, start=2):
                    row_values = [row_dict[h] for h in headers]
                    ws.append(row_values)
                    ws.row_dimensions[r_idx].height = 20
                    
                    for c_idx, h in enumerate(headers, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.font = regular_font
                        cell.border = thin_border
                        
                        if h in ['Item No (Excel)', 'Item No (PDF)', 'PDF Page', 'Qty (Excel)', 'Qty (PDF)', 'HS Code (Excel)', 'HS Code (PDF)', 'English Name Match', 'Part No Match', 'Qty Match', 'HS Code Match', 'Overall Status']:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
                    
                    # Apply colors dynamically
                    for col_idx, (match_key, cols_to_color) in enumerate([
                        ('English Name Match', [4, 5, 6]),
                        ('Part No Match', [7, 8, 9]),
                        ('Qty Match', [10, 11, 12]),
                        ('HS Code Match', [13, 14, 15])
                    ]):
                        is_match = row_dict[match_key] == "MATCH"
                        fill_to_use = match_fill if is_match else mismatch_fill
                        font_to_use = match_font if is_match else mismatch_font
                        for col in cols_to_color:
                            ws.cell(row=r_idx, column=col).fill = fill_to_use
                            ws.cell(row=r_idx, column=col).font = font_to_use
                            
                    # Overall Status
                    is_overall_match = row_dict['Overall Status'] == "MATCH"
                    status_cell = ws.cell(row=r_idx, column=16)
                    status_cell.fill = status_match_fill if is_overall_match else status_mismatch_fill
                    status_cell.font = status_match_font if is_overall_match else status_mismatch_font
                    
                for col in ws.columns:
                    col_max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or '')
                        lines = val_str.split('\n')
                        for line in lines:
                            if len(line) > col_max_len:
                                col_max_len = len(line)
                    ws.column_dimensions[col_letter].width = max(min(col_max_len + 3, 50), 10)
                
                # Save Excel workbook to bytes buffer
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_bytes = excel_buffer.getvalue()
                
                # Write CSV in memory
                csv_buffer = io.StringIO()
                writer = csv.DictWriter(csv_buffer, fieldnames=headers)
                writer.writeheader()
                writer.writerows(report_rows)
                csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
                
                # ------------------- Streamlit UI Output -------------------
                st.balloons()
                
                # Section 1: Summary Metrics
                st.subheader("📋 สรุปผลการตรวจสอบ (Audit Summary)")
                m1, m2, m3 = st.columns(3)
                with m1:
                    st.markdown(f'<div class="metric-card"><div class="metric-val">{max_len}</div><div>รายการทั้งหมด (Total)</div></div>', unsafe_allow_html=True)
                with m2:
                    st.markdown(f'<div class="metric-card" style="border-top: 5px solid #2ecc71;"><div class="metric-val" style="color: #2ecc71;">{max_len - mismatches_count}</div><div>ผ่านเกณฑ์ (Matches)</div></div>', unsafe_allow_html=True)
                with m3:
                    st.markdown(f'<div class="metric-card" style="border-top: 5px solid #e74c3c;"><div class="metric-val" style="color: #e74c3c;">{mismatches_count}</div><div>ไม่ผ่านเกณฑ์ (Mismatches)</div></div>', unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                if mismatches_count == 0:
                    st.success("🎉 ตรวจสอบเสร็จสิ้น: ข้อมูลทั้งหมดตรงกันสมบูรณ์ 100%! ไม่พบข้อผิดพลาดใดๆ")
                else:
                    st.warning(f"⚠️ ตรวจสอบเสร็จสิ้น: พบข้อมูลไม่ตรงกันจำนวน {mismatches_count} รายการ กรุณาตรวจเช็คที่ตารางหรือรายงานดาวน์โหลดด้านล่าง")
                
                # Section 2: Downloads
                st.subheader("📥 ดาวน์โหลดเอกสารรายงาน (Download Reports)")
                d1, d2 = st.columns(2)
                with d1:
                    st.download_button(
                        label="🟢 ดาวน์โหลดรายงานเปรียบเทียบ Excel (.xlsx)",
                        data=excel_bytes,
                        file_name="form_e_comparison_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with d2:
                    st.download_button(
                        label="📄 ดาวน์โหลดรายงานเปรียบเทียบ CSV (.csv)",
                        data=csv_bytes,
                        file_name="form_e_comparison_report.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                
                # Section 3: Dataframe Preview with styling
                st.subheader("🔍 ตารางแสดงผลลัพธ์การตรวจสอบ (Preview Table)")
                
                # Dynamic coloring for Streamlit dataframe view
                def style_dataframe_rows(row):
                    if row['Overall Status'] == 'MATCH':
                        return ['background-color: #E2EFDA; color: #375623'] * len(row)
                    else:
                        return ['background-color: #FCE4D6; color: #C00000; font-weight: bold'] * len(row)
                
                styled_df = df_report.style.apply(style_dataframe_rows, axis=1)
                st.dataframe(styled_df, use_container_width=True, height=500)
                
            except Exception as e:
                import traceback
                st.error("เกิดข้อผิดพลาดในการตรวจสอบเอกสาร")
                st.code(traceback.format_exc())
else:
    st.info("💡 กรุณาอัปโหลดทั้งไฟล์ Excel และ PDF ด้านบนเพื่อเริ่มกระบวนการตรวจสอบ")
