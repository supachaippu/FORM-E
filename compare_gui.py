import os
import re
import csv
import sys
import threading
import pandas as pd
import fitz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Tkinter Imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

# Helper functions for comparison
def clean_part_no(part):
    if not part or pd.isna(part):
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(part)).upper()

def clean_eng_name(name):
    if not name or pd.isna(name):
        return ""
    s = str(name).upper()
    # Remove hyphens followed by space/newline, then remove all non-alphanumeric
    s = s.replace('- ', '').replace('-\n', '').replace('-', '')
    return re.sub(r'[^A-Z0-9]', '', s)

def extract_fields_from_pdf(pdf_desc, excel_part_no):
    # Strip package prefix
    prefix_match = re.match(r'^(?:[A-Z\s\-]+ \(\d+\) PACKAGES? OF\s+)', pdf_desc, re.IGNORECASE)
    if prefix_match:
        prefix = prefix_match.group(0)
        remaining = pdf_desc[len(prefix):].strip()
    else:
        remaining = pdf_desc.strip()
        
    # Extract HS Code
    hs_code = ""
    hs_match = re.search(r'HS\s*CODE:\s*([0-9\.]+)', remaining, re.IGNORECASE)
    if hs_match:
        hs_code = hs_match.group(1).strip()
        remaining = remaining[:hs_match.start()].strip()
        
    # Extract Part Number and English Name
    clean_ex = clean_part_no(excel_part_no)
    
    alpha_indices = [i for i, c in enumerate(remaining) if c.isalnum()]
    alpha_chars = ''.join([remaining[i].upper() for i in alpha_indices])
    
    match_idx = alpha_chars.find(clean_ex)
    if match_idx != -1:
        start_in_remaining = alpha_indices[match_idx]
        end_in_remaining = alpha_indices[match_idx + len(clean_ex) - 1] + 1
        
        eng_name = remaining[:start_in_remaining].strip()
        pdf_part = remaining[start_in_remaining:end_in_remaining].strip()
        
        eng_name = re.sub(r'[\s\-\,]+$', '', eng_name)
        return eng_name, pdf_part, hs_code
    else:
        parts = re.split(r'\s{2,}', remaining)
        if len(parts) >= 2:
            return parts[0], parts[1], hs_code
        else:
            return remaining, "", hs_code

def extract_pdf_items(pdf_path, log_callback):
    doc = fitz.open(pdf_path)
    pdf_items = []
    current_item = None
    
    log_callback(f"Parsing PDF: {os.path.basename(pdf_path)} ({len(doc)} pages)...")
    
    for page_idx in range(len(doc)):
        page_num = page_idx + 1
        page = doc[page_idx]
        blocks = page.get_text("blocks")
        
        page_items = []
        page_qtys = []
        page_descs = []
        
        for b in blocks:
            x0, y0, x1, y1, text, block_no, block_type = b
            text = text.strip()
            if not text:
                continue
                
            x0 = round(x0, 1)
            y0 = round(y0, 1)
            x1 = round(x1, 1)
            y1 = round(y1, 1)
            
            # Item number (typically x0 around 65-85, y0 in the table area)
            if 65 <= x0 <= 85 and 350 <= y0 <= 600:
                lines = [l.strip() for l in text.split('\n') if l.strip()]
                for line in lines:
                    try:
                        val = int(line)
                        page_items.append((y0, val))
                    except ValueError:
                        pass
                        
            # Description column (x0 around 160-180)
            elif 160 <= x0 <= 180 and 350 <= y0 <= 600:
                page_descs.append((y0, text))
                
            # Quantity column (x0 around 420-470)
            elif 420 <= x0 <= 470 and 350 <= y0 <= 600:
                page_qtys.append((y0, text))
                
        # Sort by y coordinate
        page_items.sort(key=lambda x: x[0])
        page_qtys.sort(key=lambda x: x[0])
        page_descs.sort(key=lambda x: x[0])
        
        # Process continuation from previous page
        if page_descs:
            first_item_y = page_items[0][0] if page_items else 999.0
            first_desc_y, first_desc_text = page_descs[0]
            
            if first_desc_y < first_item_y and first_desc_y < 390:
                if page_num == 1 and first_desc_text == 'TRUCK SPARE PARTS':
                    page_descs.pop(0)
                else:
                    if current_item is not None:
                        current_item['descs'].append(first_desc_text)
                        page_descs.pop(0)
                    else:
                        page_descs.pop(0)
                        
        # Match items, descs, and qtys on this page
        for y_item, item_no in page_items:
            # Find closest qty
            matched_qty = ""
            best_qty_diff = 999.0
            best_qty_idx = -1
            for idx, (y_qty, qty_text) in enumerate(page_qtys):
                diff = abs(y_qty - y_item)
                if diff < best_qty_diff:
                    best_qty_diff = diff
                    best_qty_idx = idx
            if best_qty_idx != -1 and best_qty_diff < 10.0:
                matched_qty = page_qtys[best_qty_idx][1]
                page_qtys.pop(best_qty_idx)
                
            # Find closest desc
            matched_desc = ""
            best_desc_diff = 999.0
            best_desc_idx = -1
            for idx, (y_desc, desc_text) in enumerate(page_descs):
                diff = abs(y_desc - y_item)
                if diff < best_desc_diff:
                    best_desc_diff = diff
                    best_desc_idx = idx
            if best_desc_idx != -1 and best_desc_diff < 10.0:
                matched_desc = page_descs[best_desc_idx][1]
                page_descs.pop(best_desc_idx)
                
            current_item = {
                'item_no': item_no,
                'qty_raw': matched_qty,
                'descs': [matched_desc] if matched_desc else [],
                'page': page_num
            }
            pdf_items.append(current_item)
            
    # Post-process extracted PDF items
    processed_items = []
    for item in pdf_items:
        full_desc = " ".join(item['descs']).replace('\n', ' ')
        
        # Parse Qty
        qty_val = 0
        qty_num_match = re.search(r'(\d+)', item['qty_raw'])
        if qty_num_match:
            qty_val = int(qty_num_match.group(1))
            
        processed_items.append({
            'item_no': item['item_no'],
            'qty_raw': item['qty_raw'],
            'qty_val': qty_val,
            'full_desc': full_desc,
            'page': item['page']
        })
        
    log_callback(f"Successfully extracted {len(processed_items)} items from PDF.")
    return processed_items

def run_comparison(excel_path, pdf_path, output_dir, log_callback):
    try:
        log_callback("Reading Excel file...")
        # Load Excel items
        df_excel = pd.read_excel(excel_path, sheet_name="INVOICE FOR CNY")
        
        excel_items = []
        for idx, row in df_excel.iterrows():
            try:
                sl = int(row.iloc[0])
                part_no = str(row.iloc[5]).strip()
                ocr_no = str(row.iloc[6]).strip()
                qty = int(row.iloc[7])
                hs_code = str(row.iloc[11]).strip()
                thai_name = str(row.iloc[2]).strip()
                eng_name = str(row.iloc[4]).strip()
                
                excel_items.append({
                    'sl': sl, 'part_no': part_no, 'ocr_no': ocr_no, 'qty': qty,
                    'hs_code': hs_code, 'thai_name': thai_name, 'eng_name': eng_name
                })
            except:
                continue
                
        log_callback(f"Loaded {len(excel_items)} items from Excel sheet.")
        
        # Parse PDF items
        pdf_items = extract_pdf_items(pdf_path, log_callback)
        
        log_callback("Comparing Excel and PDF items...")
        report_rows = []
        mismatches_count = 0
        
        max_len = max(len(excel_items), len(pdf_items))
        for i in range(max_len):
            ex = excel_items[i] if i < len(excel_items) else None
            pdf = pdf_items[i] if i < len(pdf_items) else None
            
            status = "MATCH"
            mismatch_reasons = []
            
            if ex and pdf:
                pdf_eng, pdf_part, pdf_hs = extract_fields_from_pdf(pdf['full_desc'], ex['part_no'])
                
                # Check English name
                clean_ex_eng = clean_eng_name(ex['eng_name'])
                clean_pdf_eng = clean_eng_name(pdf_eng)
                name_match = (clean_ex_eng == clean_pdf_eng)
                
                # Check Part number
                clean_ex_part = clean_part_no(ex['part_no'])
                clean_pdf_part = clean_part_no(pdf_part)
                part_match = (clean_ex_part == clean_pdf_part)
                
                # Check Qty
                qty_match = (ex['qty'] == pdf['qty_val'])
                
                # Check HS code (first 6 digits)
                clean_ex_hs = re.sub(r'[^0-9]', '', ex['hs_code'])[:6]
                clean_pdf_hs = re.sub(r'[^0-9]', '', pdf_hs)[:6]
                hs_match = (clean_ex_hs == clean_pdf_hs)
                
                if not name_match:
                    status = "MISMATCH"
                    mismatch_reasons.append("English Name mismatch")
                if not part_match:
                    status = "MISMATCH"
                    mismatch_reasons.append("Part No mismatch")
                if not qty_match:
                    status = "MISMATCH"
                    mismatch_reasons.append("Qty mismatch")
                if not hs_match:
                    status = "MISMATCH"
                    mismatch_reasons.append("HS Code mismatch")
                    
                report_rows.append({
                    'Item No (Excel)': ex['sl'], 'Item No (PDF)': pdf['item_no'], 'PDF Page': pdf['page'],
                    'English Name (Excel)': ex['eng_name'], 'English Name (PDF)': pdf_eng,
                    'English Name Match': "MATCH" if name_match else "MISMATCH",
                    'Part No (Excel)': ex['part_no'], 'Part No (PDF)': pdf_part,
                    'Part No Match': "MATCH" if part_match else "MISMATCH",
                    'Qty (Excel)': ex['qty'], 'Qty (PDF)': pdf['qty_raw'],
                    'Qty Match': "MATCH" if qty_match else "MISMATCH",
                    'HS Code (Excel)': ex['hs_code'], 'HS Code (PDF)': pdf_hs,
                    'HS Code Match': "MATCH" if hs_match else "MISMATCH",
                    'Overall Status': status,
                    'Details': "; ".join(mismatch_reasons) if mismatch_reasons else "All matched"
                })
                
                if status == "MISMATCH":
                    mismatches_count += 1
                    
            elif ex:
                report_rows.append({
                    'Item No (Excel)': ex['sl'], 'Item No (PDF)': "", 'PDF Page': "",
                    'English Name (Excel)': ex['eng_name'], 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                    'Part No (Excel)': ex['part_no'], 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                    'Qty (Excel)': ex['qty'], 'Qty (PDF)': "", 'Qty Match': "MISMATCH",
                    'HS Code (Excel)': ex['hs_code'], 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                    'Overall Status': "MISMATCH", 'Details': "Item missing in PDF"
                })
                mismatches_count += 1
            elif pdf:
                report_rows.append({
                    'Item No (Excel)': "", 'Item No (PDF)': pdf['item_no'], 'PDF Page': pdf['page'],
                    'English Name (Excel)': "", 'English Name (PDF)': "", 'English Name Match': "MISMATCH",
                    'Part No (Excel)': "", 'Part No (PDF)': "", 'Part No Match': "MISMATCH",
                    'Qty (Excel)': "", 'Qty (PDF)': pdf['qty_raw'], 'Qty Match': "MISMATCH",
                    'HS Code (Excel)': "", 'HS Code (PDF)': "", 'HS Code Match': "MISMATCH",
                    'Overall Status': "MISMATCH", 'Details': "Item missing in Excel"
                })
                mismatches_count += 1
                
        # Export paths
        csv_path = os.path.join(output_dir, "form_e_comparison_report.csv")
        xlsx_path = os.path.join(output_dir, "form_e_comparison_report.xlsx")
        
        # Write CSV report
        headers = [
            'Item No (Excel)', 'Item No (PDF)', 'PDF Page', 
            'English Name (Excel)', 'English Name (PDF)', 'English Name Match',
            'Part No (Excel)', 'Part No (PDF)', 'Part No Match',
            'Qty (Excel)', 'Qty (PDF)', 'Qty Match',
            'HS Code (Excel)', 'HS Code (PDF)', 'HS Code Match',
            'Overall Status', 'Details'
        ]
        
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(report_rows)
        log_callback(f"CSV report written to: {csv_path}")

        # Write Styled Excel report
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison Report"
        ws.views.sheetView[0].showGridLines = True
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        
        regular_font = Font(name=font_family, size=11)
        
        match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        match_font = Font(name=font_family, size=11, color="375623")
        
        mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        mismatch_font = Font(name=font_family, size=11, color="C00000", bold=True)
        
        status_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        status_match_font = Font(name=font_family, size=11, color="006100", bold=True)
        status_mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_mismatch_font = Font(name=font_family, size=11, color="9C0006", bold=True)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[1].height = 28
        
        for r_idx, row_dict in enumerate(report_rows, start=2):
            row_values = [row_dict[h] for h in headers]
            ws.append(row_values)
            ws.row_dimensions[r_idx].height = 20
            
            for c_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border
                
                if h in ['Item No (Excel)', 'Item No (PDF)', 'PDF Page', 'Qty (Excel)', 'Qty (PDF)', 'HS Code (Excel)', 'HS Code (PDF)', 'English Name Match', 'Part No Match', 'Qty Match', 'HS Code Match', 'Overall Status']:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
            
            # English Name Match color
            name_match_status = row_dict['English Name Match']
            fill_to_use = match_fill if name_match_status == "MATCH" else mismatch_fill
            font_to_use = match_font if name_match_status == "MATCH" else mismatch_font
            ws.cell(row=r_idx, column=4).fill = fill_to_use
            ws.cell(row=r_idx, column=4).font = font_to_use
            ws.cell(row=r_idx, column=5).fill = fill_to_use
            ws.cell(row=r_idx, column=5).font = font_to_use
            ws.cell(row=r_idx, column=6).fill = fill_to_use
            ws.cell(row=r_idx, column=6).font = font_to_use

            # Part No Match color
            part_match_status = row_dict['Part No Match']
            fill_to_use = match_fill if part_match_status == "MATCH" else mismatch_fill
            font_to_use = match_font if part_match_status == "MATCH" else mismatch_font
            ws.cell(row=r_idx, column=7).fill = fill_to_use
            ws.cell(row=r_idx, column=7).font = font_to_use
            ws.cell(row=r_idx, column=8).fill = fill_to_use
            ws.cell(row=r_idx, column=8).font = font_to_use
            ws.cell(row=r_idx, column=9).fill = fill_to_use
            ws.cell(row=r_idx, column=9).font = font_to_use

            # Qty Match color
            qty_match_status = row_dict['Qty Match']
            fill_to_use = match_fill if qty_match_status == "MATCH" else mismatch_fill
            font_to_use = match_font if qty_match_status == "MATCH" else mismatch_font
            ws.cell(row=r_idx, column=10).fill = fill_to_use
            ws.cell(row=r_idx, column=10).font = font_to_use
            ws.cell(row=r_idx, column=11).fill = fill_to_use
            ws.cell(row=r_idx, column=11).font = font_to_use
            ws.cell(row=r_idx, column=12).fill = fill_to_use
            ws.cell(row=r_idx, column=12).font = font_to_use

            # HS Code Match color
            hs_match_status = row_dict['HS Code Match']
            fill_to_use = match_fill if hs_match_status == "MATCH" else mismatch_fill
            font_to_use = match_font if hs_match_status == "MATCH" else mismatch_font
            ws.cell(row=r_idx, column=13).fill = fill_to_use
            ws.cell(row=r_idx, column=13).font = font_to_use
            ws.cell(row=r_idx, column=14).fill = fill_to_use
            ws.cell(row=r_idx, column=14).font = font_to_use
            ws.cell(row=r_idx, column=15).fill = fill_to_use
            ws.cell(row=r_idx, column=15).font = font_to_use

            # Overall status color
            overall_status = row_dict['Overall Status']
            status_cell = ws.cell(row=r_idx, column=16)
            if overall_status == "MATCH":
                status_cell.fill = status_match_fill
                status_cell.font = status_match_font
            else:
                status_cell.fill = status_mismatch_fill
                status_cell.font = status_mismatch_font
                
        # Auto adjust column widths
        for col in ws.columns:
            col_max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > col_max_len:
                        col_max_len = len(line)
            ws.column_dimensions[col_letter].width = max(min(col_max_len + 3, 50), 10)
            
        wb.save(xlsx_path)
        log_callback(f"Excel report written to: {xlsx_path}")
        
        log_callback(f"\n--- Comparison Completed ---")
        log_callback(f"Total rows compared: {max_len}")
        log_callback(f"Mismatches found: {mismatches_count}")
        log_callback(f"All reports saved to output directory.")
        
        return True, max_len, mismatches_count, xlsx_path
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        log_callback(f"\nError running comparison:\n{err_msg}")
        return False, str(e), 0, ""

# GUI Window Class
class CompareApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FORM E - Comparative Verification Tool")
        self.root.geometry("750x600")
        self.root.minsize(700, 500)
        
        # Set ttk style
        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        # Configure custom colors
        self.style.configure(".", font=("Segoe UI", 10))
        self.style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"), foreground="#29465B")
        self.style.configure("Browse.TButton", font=("Segoe UI", 9))
        self.style.configure("Action.TButton", font=("Segoe UI", 11, "bold"), foreground="white", background="#365F91")
        self.style.map("Action.TButton", background=[("active", "#29465B")])
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header Label
        header_lbl = ttk.Label(main_frame, text="ระบบตรวจสอบและเปรียบเทียบเอกสาร FORM E", style="Header.TLabel")
        header_lbl.pack(anchor=tk.W, pady=(0, 5))
        
        subheader_lbl = ttk.Label(main_frame, text="เปรียบเทียบรายการสินค้า English Name, Part No, Qty, และ HS Code ระหว่าง Excel และ PDF (โดยไม่มีการใช้ AI)", font=("Segoe UI", 9, "italic"), foreground="#555555")
        subheader_lbl.pack(anchor=tk.W, pady=(0, 15))
        
        # Inputs Frame
        inputs_frame = ttk.LabelFrame(main_frame, text="เลือกไฟล์และโฟลเดอร์ปลายทาง (Select Files & Output)", padding="10")
        inputs_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Grid layout inside inputs frame
        inputs_frame.columnconfigure(1, weight=1)
        
        # Row 1: Excel File
        ttk.Label(inputs_frame, text="ไฟล์ Excel (Invoice):").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.excel_var = tk.StringVar()
        self.excel_entry = ttk.Entry(inputs_frame, textvariable=self.excel_var)
        self.excel_entry.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=5)
        ttk.Button(inputs_frame, text="Browse...", command=self.browse_excel, style="Browse.TButton").grid(row=0, column=2, pady=5)
        
        # Row 2: PDF File
        ttk.Label(inputs_frame, text="ไฟล์ PDF (Form E):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.pdf_var = tk.StringVar()
        self.pdf_entry = ttk.Entry(inputs_frame, textvariable=self.pdf_var)
        self.pdf_entry.grid(row=1, column=1, sticky=tk.EW, padx=5, pady=5)
        ttk.Button(inputs_frame, text="Browse...", command=self.browse_pdf, style="Browse.TButton").grid(row=1, column=2, pady=5)
        
        # Row 3: Output Folder
        ttk.Label(inputs_frame, text="โฟลเดอร์ปลายทาง:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_var = tk.StringVar()
        self.output_entry = ttk.Entry(inputs_frame, textvariable=self.output_var)
        self.output_entry.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=5)
        ttk.Button(inputs_frame, text="Browse...", command=self.browse_output, style="Browse.TButton").grid(row=2, column=2, pady=5)
        
        # Action Button Frame
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.run_btn = ttk.Button(action_frame, text="เริ่มเปรียบเทียบเอกสาร (Start Comparison)", command=self.start_thread, style="Action.TButton")
        self.run_btn.pack(fill=tk.X, ipady=5)
        
        # Console / Logs Frame
        log_frame = ttk.LabelFrame(main_frame, text="บันทึกการทำงาน (Log & Console output)", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.console = ScrolledText(log_frame, height=12, font=("Courier New", 9), background="#F8F9FA", foreground="#212529")
        self.console.pack(fill=tk.BOTH, expand=True)
        self.console.insert(tk.END, "ระบบพร้อมทำงาน...\nกรุณาเลือกไฟล์ Excel, ไฟล์ PDF และระบุโฟลเดอร์ปลายทาง จากนั้นกดปุ่ม 'เริ่มเปรียบเทียบเอกสาร'\n")
        self.console.configure(state=tk.DISABLED)
        
    def log(self, message):
        self.console.configure(state=tk.NORMAL)
        self.console.insert(tk.END, message + "\n")
        self.console.see(tk.END)
        self.console.configure(state=tk.DISABLED)
        
    def browse_excel(self):
        filename = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel (Invoice and Packing list)",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.excel_var.set(filename)
            # Default output folder to same folder as excel file if empty
            if not self.output_var.get():
                self.output_var.set(os.path.dirname(filename))
                
    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="เลือกไฟล์ PDF (FORM E)",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_var.set(filename)
            # Default output folder to same folder as PDF file if empty
            if not self.output_var.get():
                self.output_var.set(os.path.dirname(filename))
                
    def browse_output(self):
        dirname = filedialog.askdirectory(title="เลือกโฟลเดอร์ปลายทางเพื่อเซฟรายงาน")
        if dirname:
            self.output_var.set(dirname)
            
    def start_thread(self):
        # Validate inputs
        excel_path = self.excel_var.get().strip()
        pdf_path = self.pdf_var.get().strip()
        output_dir = self.output_var.get().strip()
        
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Error", "กรุณาเลือกไฟล์ Excel ที่มีอยู่จริง")
            return
        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "กรุณาเลือกไฟล์ PDF ที่มีอยู่จริง")
            return
        if not output_dir or not os.path.exists(output_dir):
            messagebox.showerror("Error", "กรุณาระบุโฟลเดอร์ปลายทางที่ถูกต้อง")
            return
            
        # Disable button while running
        self.run_btn.configure(state=tk.DISABLED)
        self.console.configure(state=tk.NORMAL)
        self.console.delete("1.0", tk.END)
        self.console.configure(state=tk.DISABLED)
        
        # Start matching thread
        t = threading.Thread(target=self.run_job, args=(excel_path, pdf_path, output_dir))
        t.daemon = True
        t.start()
        
    def run_job(self, excel_path, pdf_path, output_dir):
        self.log(f"Starting comparison job...")
        self.log(f"Excel: {excel_path}")
        self.log(f"PDF: {pdf_path}")
        self.log(f"Output folder: {output_dir}\n")
        
        success, info_or_msg, mismatches, xlsx_path = run_comparison(excel_path, pdf_path, output_dir, self.log)
        
        # Enable button
        self.root.after(0, self.finish_job, success, info_or_msg, mismatches, xlsx_path, output_dir)
        
    def finish_job(self, success, info_or_msg, mismatches, xlsx_path, output_dir):
        self.run_btn.configure(state=tk.NORMAL)
        if success:
            msg = f"เปรียบเทียบเรียบร้อย!\n\nจำนวนรายการทั้งหมด: {info_or_msg}\nพบข้อผิดพลาด (Mismatch): {mismatches} รายการ\n\nรายงานถูกบันทึกไว้ในโฟลเดอร์:\n{output_dir}"
            messagebox.showinfo("Success", msg)
            # Offer to open the folder
            if messagebox.askyesno("Open Folder", "ต้องการเปิดโฟลเดอร์ปลายทางเลยหรือไม่?"):
                self.open_folder(output_dir)
        else:
            messagebox.showerror("Error", f"เกิดข้อผิดพลาดระหว่างเปรียบเทียบ:\n{info_or_msg}")
            
    def open_folder(self, path):
        try:
            if sys.platform == 'win32':
                os.startfile(path)
            elif sys.platform == 'darwin':
                import subprocess
                subprocess.Popen(['open', path])
            else:
                import subprocess
                subprocess.Popen(['xdg-open', path])
        except Exception as e:
            self.log(f"Cannot open folder: {str(e)}")

# Application entry point
if __name__ == "__main__":
    root = tk.Tk()
    app = CompareApp(root)
    root.mainloop()
